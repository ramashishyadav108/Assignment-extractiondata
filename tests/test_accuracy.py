import pytest
import sys
from pathlib import Path
import openpyxl

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent / 'backend'))

from app.services.pdf_parser import PDFParser
from app.services.llm_extractor import LLMExtractor
from app.services.template_manager import TemplateManager
from app.services.validator import DataValidator
from app.config.settings import settings


def load_excel_data(excel_path):
    """Load data from Excel file for comparison."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Get headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    # Get data rows
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data_rows.append(row_data)
    
    return data_rows


def calculate_field_accuracy(extracted_value, expected_value):
    """Calculate accuracy for a single field."""
    if extracted_value is None and expected_value is None:
        return 1.0
    
    if extracted_value is None or expected_value is None:
        return 0.0
    
    # Convert to strings for comparison
    extracted_str = str(extracted_value).strip().lower()
    expected_str = str(expected_value).strip().lower()
    
    if extracted_str == expected_str:
        return 1.0
    
    # Check if extracted value contains expected or vice versa
    if extracted_str in expected_str or expected_str in extracted_str:
        return 0.8
    
    # Calculate string similarity (basic)
    matches = sum(1 for a, b in zip(extracted_str, expected_str) if a == b)
    max_len = max(len(extracted_str), len(expected_str))
    
    if max_len == 0:
        return 0.0
    
    return matches / max_len


def calculate_accuracy(extracted_data, expected_data):
    """
    Calculate overall accuracy by comparing extracted data with expected output.
    
    Returns:
        float: Accuracy score between 0 and 1
    """
    if not expected_data or len(expected_data) == 0:
        return 0.0
    
    expected_row = expected_data[0]  # Assuming single row for now
    
    total_fields = 0
    matching_fields = 0
    field_scores = {}
    
    for key, expected_value in expected_row.items():
        if key in extracted_data:
            total_fields += 1
            extracted_value = extracted_data[key]
            
            field_accuracy = calculate_field_accuracy(extracted_value, expected_value)
            field_scores[key] = field_accuracy
            matching_fields += field_accuracy
    
    if total_fields == 0:
        return 0.0
    
    accuracy = matching_fields / total_fields
    
    print("\n=== Field-by-Field Accuracy ===")
    for field, score in field_scores.items():
        status = "✓" if score >= 0.9 else "✗"
        print(f"{status} {field}: {score:.2%}")
    
    return accuracy


@pytest.mark.skipif(
    not settings.OPENAI_API_KEY and not settings.ANTHROPIC_API_KEY,
    reason="LLM API key not configured"
)
class TestAccuracy:
    """Accuracy tests comparing extracted output with expected output."""
    
    @pytest.fixture
    def pdf_parser(self):
        return PDFParser()
    
    @pytest.fixture
    def llm_extractor(self):
        return LLMExtractor()
    
    @pytest.fixture
    def template_manager(self):
        return TemplateManager(settings.TEMPLATE_DIR)
    
    @pytest.fixture
    def validator(self):
        return DataValidator()
    
    def test_best_practices_fund_accuracy(
        self, pdf_parser, llm_extractor, template_manager, validator
    ):
        """Test extraction accuracy for Best Practices Fund II."""
        pdf_path = Path('examples/sample_pdfs/Best-Practices-Fund II.pdf')
        expected_path = Path('examples/output/Best-Practices-Fund II - output.xlsx')
        
        if not pdf_path.exists():
            pytest.skip(f"Sample PDF not found: {pdf_path}")
        
        if not expected_path.exists():
            pytest.skip(f"Expected output not found: {expected_path}")
        
        # Extract text from PDF
        pdf_content = pdf_parser.extract_text_from_pdf(pdf_path)
        
        # Get template
        template = template_manager.get_template('template_1')
        
        # Extract data with LLM
        result = llm_extractor.extract_with_llm(
            pdf_content['full_text'],
            template
        )
        
        extracted_data = result['data']
        
        # Validate extraction
        is_valid, errors, warnings = validator.validate_extraction(
            extracted_data,
            pdf_content['full_text'],
            template
        )
        
        print(f"\nValidation: {'PASSED' if is_valid else 'FAILED'}")
        if errors:
            print("Errors:", errors)
        if warnings:
            print("Warnings:", warnings)
        
        # Load expected output
        expected_data = load_excel_data(expected_path)
        
        # Calculate accuracy
        accuracy = calculate_accuracy(extracted_data, expected_data)
        
        print(f"\n=== Overall Accuracy: {accuracy:.2%} ===")
        
        # Assert accuracy is above threshold
        assert accuracy >= 0.90, f"Accuracy {accuracy:.2%} below 90% threshold"
    
    def test_linolex_fund_accuracy(
        self, pdf_parser, llm_extractor, template_manager, validator
    ):
        """Test extraction accuracy for Linolex Fund LP."""
        pdf_path = Path('examples/sample_pdfs/Linolex Fund LP (1).pdf')
        expected_path = Path('examples/output/Linolex Fund LP_Extracted_Fund_Data (5) (1).xlsx')
        
        if not pdf_path.exists():
            pytest.skip(f"Sample PDF not found: {pdf_path}")
        
        if not expected_path.exists():
            pytest.skip(f"Expected output not found: {expected_path}")
        
        # Extract and test
        pdf_content = pdf_parser.extract_text_from_pdf(pdf_path)
        template = template_manager.get_template('template_1')
        
        result = llm_extractor.extract_with_llm(
            pdf_content['full_text'],
            template
        )
        
        extracted_data = result['data']
        expected_data = load_excel_data(expected_path)
        accuracy = calculate_accuracy(extracted_data, expected_data)
        
        print(f"\n=== Overall Accuracy: {accuracy:.2%} ===")
        assert accuracy >= 0.90, f"Accuracy {accuracy:.2%} below 90% threshold"
    
    def test_horizon_capital_accuracy(
        self, pdf_parser, llm_extractor, template_manager, validator
    ):
        """Test extraction accuracy for Horizon Capital."""
        pdf_path = Path('examples/sample_pdfs/Horizon Capital.pdf')
        expected_path = Path('examples/output/Horizon Capital_Extracted_Fund_Data (5).xlsx')
        
        if not pdf_path.exists():
            pytest.skip(f"Sample PDF not found: {pdf_path}")
        
        if not expected_path.exists():
            pytest.skip(f"Expected output not found: {expected_path}")
        
        # Extract and test
        pdf_content = pdf_parser.extract_text_from_pdf(pdf_path)
        template = template_manager.get_template('template_1')
        
        result = llm_extractor.extract_with_llm(
            pdf_content['full_text'],
            template
        )
        
        extracted_data = result['data']
        expected_data = load_excel_data(expected_path)
        accuracy = calculate_accuracy(extracted_data, expected_data)
        
        print(f"\n=== Overall Accuracy: {accuracy:.2%} ===")
        assert accuracy >= 0.90, f"Accuracy {accuracy:.2%} below 90% threshold"


if __name__ == '__main__':
    pytest.main([__file__, '-v', '-s'])
