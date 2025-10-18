import logging
from pathlib import Path
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


class EnhancedExcelGenerator:
    """Generate Excel files in vertical Field|Value format with multiple sheets."""

    def __init__(self):
        self.logger = logger

    def generate_portfolio_summary_excel(
        self,
        extracted_data: Dict[str, str],
        output_path: Path,
        fund_name: str = "Fund Data"
    ):
        """
        Generate Excel file with Portfolio Summary in Field | Value format.

        Args:
            extracted_data: Dictionary of field_name -> "value|source" strings
            output_path: Path to save Excel file
            fund_name: Name of the fund for labeling
        """
        try:
            wb = Workbook()

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # Create Portfolio Summary sheet
            ws = wb.create_sheet("Portfolio Summary", 0)

            # Write header row
            ws['A1'] = "Field"
            ws['B1'] = "Value"

            # Style header
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="left", vertical="center")

            for cell in [ws['A1'], ws['B1']]:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Write data rows
            row = 2
            for field_name, value_with_source in extracted_data.items():
                ws[f'A{row}'] = field_name
                ws[f'B{row}'] = value_with_source
                row += 1

            # Format columns
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 80

            # Apply borders and alignment to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Data rows
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Create placeholder sheets for other sections
            placeholder_sheets = [
                "Schedule of Investments",
                "Statement of Operations",
                "Statement of Cashflows",
                "PCAP Statement",
                "Portfolio Company profile",
                "Portfolio Company Financials",
                "Footnotes"
            ]

            for sheet_name in placeholder_sheets:
                ws_placeholder = wb.create_sheet(sheet_name)
                ws_placeholder['A1'] = "Data not yet extracted"
                ws_placeholder['A1'].font = Font(italic=True, color="666666")

            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Enhanced Excel generated: {output_path}")

        except Exception as e:
            self.logger.error(f"Error generating enhanced Excel: {str(e)}")
            raise
